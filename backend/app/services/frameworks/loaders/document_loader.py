"""Loader for frameworks uploaded from documents (XLSX, CSV, DOCX, PDF).

Supports:
- XLSX/CSV: Auto-detect or user-mapped columns for code, name, description, parent
- PDF/DOCX: AI-powered extraction of requirements into a structured hierarchy
"""

import csv
import io
import json
from pathlib import Path
from typing import Optional, Any

from app.core.config import settings
from app.models.unified_framework import FrameworkType
from app.services.frameworks.loaders.base_loader import (
    BaseFrameworkLoader,
    FrameworkData,
    RequirementData,
)


# Common column name patterns for auto-detection
COLUMN_PATTERNS = {
    "code": ["code", "id", "identifier", "ref", "reference", "number", "no", "control_id", "req_id"],
    "name": ["name", "title", "control", "requirement", "control_name", "req_name"],
    "description": ["description", "desc", "detail", "details", "text", "body", "summary"],
    "parent": ["parent", "parent_code", "parent_id", "category", "section", "group", "domain"],
    "guidance": ["guidance", "implementation", "notes", "guidance_text", "implementation_guidance"],
}


class DocumentFrameworkLoader(BaseFrameworkLoader):
    """Loader that parses framework requirements from uploaded documents."""

    def __init__(
        self,
        requirements: list[dict],
        framework_code: str,
        framework_name: str,
        framework_version: str = "1.0",
        description: Optional[str] = None,
        hierarchy_labels: Optional[list[str]] = None,
    ):
        self._requirements = requirements
        self._code = framework_code
        self._name = framework_name
        self._version = framework_version
        self._description = description
        self._hierarchy_labels = hierarchy_labels

    def load(self, db, force_reload: bool = False):
        """Override to set is_builtin=False for uploaded frameworks."""
        import uuid as _uuid
        from app.models.unified_framework import Framework, FrameworkRequirement

        framework_data = self.get_framework_data()

        existing = (
            db.query(Framework)
            .filter(Framework.code == framework_data.code)
            .first()
        )

        if existing and not force_reload:
            return existing

        if existing and force_reload:
            db.query(FrameworkRequirement).filter(
                FrameworkRequirement.framework_id == existing.id
            ).delete()
            db.delete(existing)
            db.flush()

        framework = Framework(
            id=_uuid.uuid4(),
            code=framework_data.code,
            name=framework_data.name,
            version=framework_data.version,
            description=framework_data.description,
            framework_type=framework_data.framework_type.value,
            hierarchy_levels=framework_data.hierarchy_levels,
            hierarchy_labels=framework_data.hierarchy_labels,
            extra_metadata=framework_data.metadata,
            is_active=True,
            is_builtin=False,
        )
        db.add(framework)
        db.flush()

        self._load_requirements(
            db=db,
            framework_id=framework.id,
            requirements=framework_data.requirements,
            parent_id=None,
            level=0,
        )

        db.commit()
        return framework

    def get_framework_data(self) -> FrameworkData:
        """Build FrameworkData from parsed requirements."""
        req_data = self._build_hierarchy(self._requirements)
        max_level = self._get_max_level(req_data)

        return FrameworkData(
            code=self._code,
            name=self._name,
            version=self._version,
            description=self._description,
            framework_type=FrameworkType.CUSTOM,
            hierarchy_levels=max_level + 1,
            hierarchy_labels=self._hierarchy_labels,
            requirements=req_data,
        )

    def _build_hierarchy(self, requirements: list[dict]) -> list[RequirementData]:
        """Build a hierarchy tree from flat requirement dicts."""
        # Index by code for parent lookups
        by_code: dict[str, dict] = {}
        for req in requirements:
            code = req.get("code", "").strip()
            if code:
                by_code[code] = req

        # Find roots (no parent or parent not found)
        roots = []
        children_map: dict[str, list[dict]] = {}

        for req in requirements:
            parent = req.get("parent", "").strip()
            if not parent or parent not in by_code:
                roots.append(req)
            else:
                children_map.setdefault(parent, []).append(req)

        def build_tree(items: list[dict], level: int) -> list[RequirementData]:
            result = []
            for idx, item in enumerate(items):
                code = item.get("code", f"REQ-{level}-{idx}")
                children = children_map.get(code, [])
                child_data = build_tree(children, level + 1)

                result.append(RequirementData(
                    code=code,
                    name=item.get("name", code),
                    description=item.get("description"),
                    guidance=item.get("guidance"),
                    level=level,
                    is_assessable=len(child_data) == 0,
                    display_order=idx,
                    children=child_data,
                ))
            return result

        return build_tree(roots, 0)

    def _get_max_level(self, reqs: list[RequirementData], current: int = 0) -> int:
        if not reqs:
            return current
        return max(
            self._get_max_level(req.children, req.level)
            for req in reqs
        )

    @classmethod
    def parse_spreadsheet(
        cls,
        file_content: bytes,
        filename: str,
        column_mapping: Optional[dict[str, str]] = None,
    ) -> list[dict]:
        """Parse requirements from CSV or XLSX file content.

        Args:
            file_content: Raw file bytes
            filename: Original filename (used for format detection)
            column_mapping: Optional explicit column mapping
                e.g. {"code": "Control ID", "name": "Control Name"}

        Returns:
            List of requirement dicts with keys: code, name, description, parent, guidance
        """
        ext = Path(filename).suffix.lower()

        if ext == ".csv":
            return cls._parse_csv(file_content, column_mapping)
        elif ext in (".xlsx", ".xls"):
            return cls._parse_xlsx(file_content, column_mapping)
        else:
            raise ValueError(f"Unsupported spreadsheet format: {ext}")

    @classmethod
    def _parse_csv(
        cls,
        content: bytes,
        column_mapping: Optional[dict[str, str]] = None,
    ) -> list[dict]:
        """Parse CSV content into requirement dicts."""
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        headers = reader.fieldnames or []

        mapping = column_mapping or cls._auto_detect_columns(headers)

        requirements = []
        for row in reader:
            req = cls._map_row(row, mapping)
            if req.get("code") or req.get("name"):
                requirements.append(req)

        return requirements

    @classmethod
    def _parse_xlsx(
        cls,
        content: bytes,
        column_mapping: Optional[dict[str, str]] = None,
    ) -> list[dict]:
        """Parse XLSX content into requirement dicts."""
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl is required for XLSX parsing. Install with: pip install openpyxl")

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
        mapping = column_mapping or cls._auto_detect_columns(headers)

        requirements = []
        for row_values in rows[1:]:
            row = {headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row_values) if i < len(headers)}
            req = cls._map_row(row, mapping)
            if req.get("code") or req.get("name"):
                requirements.append(req)

        return requirements

    @classmethod
    def _auto_detect_columns(cls, headers: list[str]) -> dict[str, str]:
        """Auto-detect column mapping from headers."""
        mapping = {}
        headers_lower = {h: h.lower().strip().replace(" ", "_") for h in headers}

        for field, patterns in COLUMN_PATTERNS.items():
            for header, header_lower in headers_lower.items():
                if header_lower in patterns or any(p in header_lower for p in patterns):
                    mapping[field] = header
                    break

        # If no code column found, use first column
        if "code" not in mapping and headers:
            mapping["code"] = headers[0]
        # If no name column found, use second column
        if "name" not in mapping and len(headers) > 1:
            mapping["name"] = headers[1]

        return mapping

    @classmethod
    def _map_row(cls, row: dict, mapping: dict[str, str]) -> dict:
        """Map a row using the column mapping."""
        return {
            "code": row.get(mapping.get("code", ""), "").strip(),
            "name": row.get(mapping.get("name", ""), "").strip(),
            "description": row.get(mapping.get("description", ""), "").strip(),
            "parent": row.get(mapping.get("parent", ""), "").strip(),
            "guidance": row.get(mapping.get("guidance", ""), "").strip(),
        }

    @classmethod
    def detect_columns(cls, file_content: bytes, filename: str) -> dict[str, Any]:
        """Detect available columns and auto-mapping for a file.

        Returns:
            Dict with headers and suggested mapping
        """
        ext = Path(filename).suffix.lower()

        if ext == ".csv":
            text = file_content.decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(text))
            headers = list(reader.fieldnames or [])
        elif ext in (".xlsx", ".xls"):
            try:
                import openpyxl
            except ImportError:
                raise ImportError("openpyxl required")
            wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
            ws = wb.active
            first_row = next(ws.iter_rows(max_row=1, values_only=True), None)
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(first_row or [])]
        else:
            headers = []

        mapping = cls._auto_detect_columns(headers)

        return {
            "headers": headers,
            "suggested_mapping": mapping,
        }

    @classmethod
    async def parse_document_with_ai(
        cls,
        file_content: bytes,
        filename: str,
    ) -> list[dict]:
        """Parse requirements from PDF/DOCX using AI.

        Args:
            file_content: Raw file bytes
            filename: Original filename

        Returns:
            List of requirement dicts
        """
        ext = Path(filename).suffix.lower()

        # Extract text
        if ext == ".pdf":
            text = cls._extract_pdf_text(file_content)
        elif ext in (".docx", ".doc"):
            text = cls._extract_docx_text(file_content)
        elif ext in (".txt", ".md"):
            text = file_content.decode("utf-8", errors="replace")
        else:
            raise ValueError(f"Unsupported document format: {ext}")

        if not text.strip():
            raise ValueError("No text could be extracted from the document")

        # Truncate very long documents
        max_chars = 50000
        if len(text) > max_chars:
            text = text[:max_chars] + "\n\n[Document truncated...]"

        # Use AI to parse requirements
        return await cls._ai_parse_requirements(text, filename)

    @classmethod
    def _extract_pdf_text(cls, content: bytes) -> str:
        """Extract text from a PDF file."""
        try:
            import PyPDF2
        except ImportError:
            raise ImportError("PyPDF2 is required for PDF parsing. Install with: pip install PyPDF2")

        reader = PyPDF2.PdfReader(io.BytesIO(content))
        pages = []
        for page in reader.pages:
            text = page.extract_text()
            if text:
                pages.append(text)
        return "\n\n".join(pages)

    @classmethod
    def _extract_docx_text(cls, content: bytes) -> str:
        """Extract text from a DOCX file."""
        try:
            import docx
        except ImportError:
            raise ImportError("python-docx is required for DOCX parsing. Install with: pip install python-docx")

        doc = docx.Document(io.BytesIO(content))
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    @classmethod
    async def _ai_parse_requirements(cls, text: str, filename: str) -> list[dict]:
        """Use Claude to extract structured requirements from text."""
        if not settings.anthropic_api_key:
            raise ValueError("ANTHROPIC_API_KEY required for AI document parsing")

        from anthropic import Anthropic

        client = Anthropic(api_key=settings.anthropic_api_key)

        prompt = f"""Extract compliance requirements from this document into a structured format.

DOCUMENT (from file: {filename}):
---
{text}
---

Parse the document and output a JSON array of requirement objects. Each requirement should have:
- "code": A short identifier code (e.g., "1.1", "AC-01", "REQ-001")
- "name": A brief name/title for the requirement
- "description": The full requirement description
- "parent": The parent requirement's code (empty string for top-level items)
- "guidance": Any implementation guidance or notes (empty string if none)

Preserve the document's hierarchy. Top-level sections/categories should have parent="" and
their sub-requirements should reference the parent code.

If the document doesn't have explicit codes, create logical ones based on section numbering.

Respond ONLY with the JSON array, no other text."""

        response = client.messages.create(
            model=settings.ai_model,
            max_tokens=settings.ai_max_tokens,
            temperature=0.1,
            messages=[{"role": "user", "content": prompt}],
        )

        content = response.content[0].text.strip()
        # Handle markdown code blocks
        if content.startswith("```"):
            content = content.split("```")[1]
            if content.startswith("json"):
                content = content[4:]
            content = content.strip()

        requirements = json.loads(content)

        if not isinstance(requirements, list):
            raise ValueError("AI did not return a valid requirements list")

        return requirements
